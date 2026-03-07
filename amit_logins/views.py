from django.shortcuts import render, redirect,get_object_or_404
from datetime import datetime, date
import pandas as pd
from django.http import HttpResponse
from django.contrib.auth import login, authenticate, logout
from django.views.decorators.cache import cache_control,never_cache
from .forms import UserRegisterForm, UserLoginForm
from django.contrib.auth.decorators import login_required
from .models import *
from .forms import *
from django.core.exceptions import ObjectDoesNotExist
from openpyxl.styles import NamedStyle
from django.contrib import messages
import xlwt
from itertools import chain
from collections import defaultdict
from django.db.models import Sum,F,Q, Prefetch
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.forms import modelformset_factory
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.utils.timezone import now, localtime
from django.core.mail import send_mail
from django.conf import settings
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os
from django.utils.dateparse import parse_date

def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False  
            user.save()
            return redirect('login')
    else:
        form = UserRegisterForm()
    return render(request, 'logins/register.html', {'form': form})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def user_login(request):
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('dashboard')
    else:
        form = UserLoginForm()
    return render(request, 'logins/login.html', {'form': form})

@login_required
def approve_user(request, user_id):
    if not request.user.is_superuser:
        return redirect('login')

    user = get_object_or_404(User, id=user_id)
    user.is_approved = True  
    user.is_active = True  
    user.save()

    return redirect('admin_dashboard')

@login_required
def reject_user(request, user_id):
    if not request.user.is_superuser:
        return redirect('login')

    user = get_object_or_404(User, id=user_id)
    user.delete()  
    return redirect('admin_dashboard')

@login_required
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.error(request, f"User {user.username} has been deleted.")
    return redirect('admin_dashboard')

def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        user_form = UserEditForm(request.POST, instance=user)
        password_form = UserPasswordEditForm(user, request.POST)

        if user_form.is_valid() and password_form.is_valid():
            user_form.save()
            password_form.save()
            return redirect('admin_dashboard')  
    else:
        user_form = UserEditForm(instance=user)
        password_form = UserPasswordEditForm(user)

    return render(request, 'logins/edit_user.html', {
        'user_form': user_form,
        'password_form': password_form,
        'user': user
    })

@login_required(login_url='login') 
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dashboard(request):
    if request.user.is_superuser:   
        login_time = localtime(now()).strftime('%Y-%m-%d %I:%M:%S %p')        
        send_mail(
            subject='Admin Login Notification (athithmithra.org)',
            message=f"Superuser {request.user.username} logged in at {login_time}.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=['abtechchennai@gmail.com'],  
            fail_silently=False,
        )
        return redirect('admin_dashboard')      
    context = {
        'branch': request.user.branch,
        'department': request.user.department
    }    
    if request.user.department == 'ADMIN-MANAGEMENT WORK TEAM':
        template_name = f"logins/incharge_dashboard_{request.user.branch.lower()}.html"
    else:
        template_name = f"logins/employee_dashboard_{request.user.branch.lower()}.html"

    return render(request, template_name, context)
def user_logout(request):
    logout(request)
    return redirect('login')

def admin_dashboard(request):
    user_registrations = User.objects.all().order_by('username')
    BRANCHES = ['NAGERCOIL']
    work_status_records = WorkStatus.objects.all()
    DEPARTMENT_CHOICES = [
    
    'hardware project development team',
    
    'Admin team',
]
    context = {
        'user_registrations': user_registrations,
        'departments': DEPARTMENT_CHOICES,
        'branches': BRANCHES,
    }
    return render(request, 'logins/admin_dashboard.html', context)

@login_required(login_url='login')
def workstatus_branch_list(request):
    branches = ['NAGERCOIL']
    return render(request, 'workstatus/branch_list.html', {'branches': branches})
@login_required(login_url='login')
def workstatus_department_list(request, branch):
    user = request.user

    all_departments = [
        # 'JOURNAL TEAM',
        # 'RESEARCH WORK DEVELOPMENT TEAM',
        # 'SOFTWARE PROJECT DEVELOPMENT TEAM',
        'HARDWARE PROJECT DEVELOPMENT TEAM',
        # 'TECHNICAL TEAM',
        'ADMIN TEAM',
    ]

    if user.is_superuser:
        # ✅ Superuser sees all departments
        departments = all_departments

    elif getattr(user, "is_team_leader", False):
        # ✅ Team Leader sees only their own department
        departments = [user.department]

    else:
        # ✅ Employees have no access here
        return redirect("dashboard")

    return render(request, "workstatus/department_list.html", {
        "branch": branch,
        "departments": departments,
    })
@login_required(login_url='login')
def workstatus_records_view(request, branch, department):
    query = request.GET.get('q', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    groups = WorkStatusGroup.objects.filter(
        branch=branch,
        department=department
    ).prefetch_related('statuses').select_related('user')

    # Search filter
    if query:
        groups = groups.filter(
            Q(user__username__icontains=query) |
            Q(statuses__WORK_CODE__icontains=query) |
            Q(S_NO__icontains=query)
        ).distinct()

    # Date range filter
    if from_date and to_date:
        parsed_from_date = parse_date(from_date)
        parsed_to_date = parse_date(to_date)
        if parsed_from_date and parsed_to_date:
            groups = groups.filter(DATE__range=[parsed_from_date, parsed_to_date])

    elif from_date:  # Only from_date provided
        parsed_from_date = parse_date(from_date)
        if parsed_from_date:
            groups = groups.filter(DATE__gte=parsed_from_date)

    elif to_date:  # Only to_date provided
        parsed_to_date = parse_date(to_date)
        if parsed_to_date:
            groups = groups.filter(DATE__lte=parsed_to_date)

    groups = groups.order_by('DATE', 'user__username')

    return render(request, 'workstatus/workstatus_records.html', {
        'branch': branch,
        'department': department,
        'groups': groups,
        'from_date': from_date,
        'to_date': to_date,
        'query': query,
    })



# EMPLOYEE DASHBOARD VIEWS
@login_required
def work_status_group_list(request):
    groups = WorkStatusGroup.objects.filter(user=request.user).prefetch_related('statuses').order_by('-DATE')
    
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]

    return render(request, 'workstatus/work_status_group_list.html', {
        'groups': groups,
        'years': years,
    })

@login_required
def create_work_status_group(request):
    if request.method == 'POST':
        form = WorkStatusGroupForm(request.POST)
        if form.is_valid():
            group = form.save(commit=False)
            group.user = request.user
            group.branch = request.user.branch
            group.department = request.user.department
            group.save()
            return redirect('work_status_group_list')
    else:
        form = WorkStatusGroupForm()

    return render(request, 'workstatus/add_edit_work_status_group.html', {
        'form': form,
        'action': 'Add'
    })

@login_required
def edit_work_status_group(request, group_id):
    group = get_object_or_404(WorkStatusGroup, id=group_id)

    if request.method == 'POST':
        form = WorkStatusGroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            return redirect('work_status_group_list')
    else:
        form = WorkStatusGroupForm(instance=group)

    return render(request, 'workstatus/add_edit_work_status_group.html', {
        'form': form,
        'action': 'Edit'
    })

@login_required
def delete_work_status_group(request, group_id):
    group = get_object_or_404(WorkStatusGroup, id=group_id)
    group.delete()
    return redirect('work_status_group_list')

@login_required
def add_work_status(request, group_id):
    group = get_object_or_404(WorkStatusGroup, id=group_id)

    if request.method == 'POST':
        form = WorkStatusForm(request.POST)
        if form.is_valid():
            status = form.save(commit=False)
            status.group = group

            # Replace blank date fields with '-'
            if not status.STARTING_DATE:
                status.STARTING_DATE = None
            if not status.ENDING_DATE:
                status.ENDING_DATE = None

            status.save()
            return redirect('work_status_group_list')
    else:
        form = WorkStatusForm()

    return render(request, 'workstatus/add_edit_work_status.html', {
        'form': form,
        'group': group,
        'action': 'Add'
    })

@login_required
def edit_work_status(request, status_id):
    status = get_object_or_404(WorkStatus, id=status_id)
    group = status.group

    if request.method == 'POST':
        form = WorkStatusForm(request.POST, instance=status)
        if form.is_valid():
            status = form.save(commit=False)
            status.group = group
            # Replace blank date fields with '-'
            if not status.STARTING_DATE:
                status.STARTING_DATE = None
            if not status.ENDING_DATE:
                status.ENDING_DATE = None

            status.save()
            return redirect('work_status_group_list')
    else:
        form = WorkStatusForm(instance=status)

    return render(request, 'workstatus/add_edit_work_status.html', {
        'form': form,
        'group': group,
        'action': 'Edit'
    })


@login_required
def delete_work_status(request, status_id):
    status = get_object_or_404(WorkStatus, id=status_id)
    status.delete()
    return redirect('work_status_group_list')


@login_required
def download_work_status_report(request):
    month = request.GET.get('month')
    year = request.GET.get('year')

    groups = WorkStatusGroup.objects.filter(user=request.user).prefetch_related('statuses')
    if month and year:
        groups = groups.filter(DATE__year=year, DATE__month=month)
    groups = groups.order_by('DATE')

    if not groups.exists():
        return HttpResponse("No records found.", status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Status Report"

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    data_font = Font(name="Bookman Old Style")
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30
    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "logo.png")
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width, logo.height = 90, 50
        ws.add_image(logo, "A1")

    ws.merge_cells('A1:G2')
    ws['A1'].value = "ATHITH MITHRA INDUSTRIAL PVT LTD"
    ws['A1'].font = Font(bold=True, color="FF0000", size=18, name="Bookman Old Style")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A3:B3')
    ws['A3'] = "NAME"
    ws['A3'].font = Font(bold=True, name="Bookman Old Style")
    ws.merge_cells('C3:D3')
    ws['C3'] = getattr(request.user, 'employee_name', '')
    ws['C3'].font = Font(bold=True, name="Bookman Old Style")

    ws.merge_cells('A4:B4')
    ws['A4'] = "EMPLOYEE ID"
    ws['A4'].font = Font(bold=True, name="Bookman Old Style")
    ws.merge_cells('C4:D4')
    ws['C4'] = request.user.username
    ws['C4'].font = Font(bold=True, name="Bookman Old Style")

    ws['E3'] = "BRANCH"
    ws['E3'].font = Font(bold=True, name="Bookman Old Style")
    ws.merge_cells('F3:G3')
    ws['F3'] = getattr(request.user, 'branch', '')
    ws['F3'].font = Font(bold=True, name="Bookman Old Style")

    ws['E4'] = "DEPARTMENT"
    ws['E4'].font = Font(bold=True, name="Bookman Old Style")
    ws.merge_cells('F4:G4')
    ws['F4'] = getattr(request.user, 'department', '')
    ws['F4'].font = Font(bold=True, name="Bookman Old Style")
    for row in range(3, 5):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border

    
    ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
    ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)
    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=7)

    ws['A5'] = "S.NO"
    ws['B5'] = "DATE"
    ws['C5'] = "WORK STATUS"

    for cell_ref in ['A5', 'B5', 'C5']:
        cell = ws[cell_ref]
        cell.font = Font(bold=True, name="Bookman Old Style")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.border = border

    sub_headers = ["WORK CODE", "WORK DETAILS", "STARTING DATE", "ENDING DATE", "STATUS"]
    for idx, header in enumerate(sub_headers, start=3):
        cell = ws.cell(row=6, column=idx, value=header)
        cell.font = Font(bold=True, name="Bookman Old Style")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.border = border

    ws.row_dimensions[5].height = 25
    ws.row_dimensions[6].height = 25
    for row in [5, 6]:
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border

    current_row = 7
    for i, group in enumerate(groups, start=1):
        statuses = list(group.statuses.all())
        first_row = current_row

        for status in statuses:
            ws.cell(row=current_row, column=1, value=group.S_NO)
            ws.cell(row=current_row, column=2, value=group.DATE.strftime('%d-%m-%Y') if group.DATE else '')
            ws.cell(row=current_row, column=3, value=status.WORK_CODE)
            ws.cell(row=current_row, column=4, value=status.WORK_DETAILS)
            ws.cell(row=current_row, column=5, value=status.STARTING_DATE.strftime('%d-%m-%Y') if status.STARTING_DATE else '-')
            ws.cell(row=current_row, column=6, value=status.ENDING_DATE.strftime('%d-%m-%Y') if status.ENDING_DATE else '-')
            ws.cell(row=current_row, column=7, value=status.WORK_STATUS)

            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True) if col != 4 else Alignment(horizontal="left", vertical="top", wrap_text=True)

            current_row += 1

        last_row = current_row - 1
        if last_row > first_row:
            ws.merge_cells(start_row=first_row, start_column=1, end_row=last_row, end_column=1)
            ws.merge_cells(start_row=first_row, start_column=2, end_row=last_row, end_column=2)

    col_widths = [10, 15, 20, 40, 23, 23, 38]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"WORK_STATUS_REPORT_{request.user.username}_{month or 'All'}_{year or 'All'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

payment_model_map = {
    "NAGERCOIL": NagercoilPaymentVoucher,
    "TIRUNELVELI": TirunelveliPaymentVoucher,
    "PUDUKOTTAI": PudukottaiPaymentVoucher,
    "CHENNAI": ChennaiPaymentVoucher,
}

payment_form_map = {
    "NAGERCOIL": NagercoilPaymentVoucherForm,
    "TIRUNELVELI": TirunelveliPaymentVoucherForm,
    "PUDUKOTTAI": PudukottaiPaymentVoucherForm,
    "CHENNAI": ChennaiPaymentVoucherForm,
}


@login_required(login_url='login')
def payment_voucher_report(request, branch):
    if branch not in payment_model_map:
        return redirect('dashboard')

    records = payment_model_map[branch].objects.all().order_by('-DATE','-S_NO' )
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]

    return render(request, "logins/payment_voucher_report.html", {
        "branch": branch,
        "records": records,
        "years": years,
    })


@login_required(login_url='login')
def add_payment_voucher(request, branch):
    form_class = payment_form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('payment_voucher_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_payment_voucher.html", {
        "form": form,
        "branch_name": branch
    })


@login_required(login_url='login')
def edit_payment_voucher(request, branch, record_id):
    if branch not in payment_model_map:
        return redirect('dashboard')

    model = payment_model_map[branch]
    form_class = payment_form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('payment_voucher_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_payment_voucher.html", {
        "form": form,
        "branch": branch
    })


@login_required(login_url='login')
def delete_payment_voucher(request, branch, record_id):
    if branch not in payment_model_map:
        return redirect('dashboard')

    model = payment_model_map[branch]
    record = get_object_or_404(model, id=record_id)

    
    record.delete()
    messages.success(request, "Payment voucher deleted successfully.")
    return redirect('payment_voucher_report', branch=branch)


def download_payment_voucher_report(request, branch):
    model_map = {
        'NAGERCOIL': NagercoilPaymentVoucher,
        'TIRUNELVELI': TirunelveliPaymentVoucher,
        'PUDUKOTTAI': PudukottaiPaymentVoucher,
        'CHENNAI': ChennaiPaymentVoucher
    }
    model = model_map.get(branch)
    
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    
    return download_report_payment_voucher(request, model, branch, 'PAYMENT_VOUCHER_REPORT')


@login_required
def download_report_payment_voucher(request, model, branch, filename):
    month = request.GET.get('month')
    year = request.GET.get('year')

    if not month or not year:
        return HttpResponse("Month and Year parameters are required.", status=400)

    records = model.objects.filter(DATE__year=year, DATE__month=month).order_by('S_NO', 'DATE')
    if not records.exists():
        return HttpResponse("No records found for the selected month and year.", status=404)

    # === Prepare Data ===
    data = []
    total_online = 0
    total_cash = 0
    for record in records:
        online = record.ONLINE_PAYMENT or 0
        cash = record.CASH_PAYMENT or 0
        total_online += online
        total_cash += cash
        data.append([
            record.S_NO,
            record.DATE.strftime('%d.%m.%Y'), 
            record.VC_NO,
            record.PURPOSE,
            online,
            cash
        ])

    # Branch and Month mapping
    branch_map = {
        'NAGERCOIL': 'NGL', 'TIRUNELVELI': 'TVL', 'CHENNAI': 'TBM', 'PUDUKOTTAI': 'PDKT', 'VIJAYAWADA': 'VJW'
    }
    month_names = {
        '1': 'JANUARY', '2': 'FEBRUARY', '3': 'MARCH', '4': 'APRIL',
        '5': 'MAY', '6': 'JUNE', '7': 'JULY', '8': 'AUGUST',
        '9': 'SEPTEMBER', '10': 'OCTOBER', '11': 'NOVEMBER', '12': 'DECEMBER'
    }
    m_key = str(int(month)) if month and month.isdigit() else str(month)
    m_name = month_names.get(m_key, 'All')
    b_code = branch_map.get(branch.upper(), branch[:3].upper())

    # === Workbook Setup ===
    wb = Workbook()
    ws = wb.active
    m_short = m_name[:3].upper()
    y_short = str(year)[-2:]
    ws.title = f"AMIT_{b_code}_PV {m_short}'{y_short}"

    # === Styles ===
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    title_font = Font(bold=True, name="Bookman Old Style", size=18)
    header_font = Font(bold=True, name="Bookman Old Style", size=11)
    data_font = Font(name="Bookman Old Style", size=10)
    data_font_bold = Font(bold=True, name="Bookman Old Style", size=11)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align_v = Alignment(horizontal="right", vertical="center")


    # === LOGO + TITLE ===
    # === LOGO + TITLE ===
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 25
    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "logo.png")
    if os.path.exists(logo_path):
        logo = XLImage(logo_path); logo.width, logo.height = 130, 85
        ws.add_image(logo, "A1")
    ws.merge_cells('A1:B4')
    ws.merge_cells('C1:G2')
    ws['C1'].value = "ATHITH MITHRA INDUSTRIAL PVT LTD"; ws['C1'].font = title_font; ws['C1'].alignment = center_align
    ws.merge_cells('C3:G4')
    ws['C3'].value = "PAYMENT VOUCHER DETAILS"; ws['C3'].font = Font(bold=True, name="Bookman Old Style", size=14); ws['C3'].alignment = center_align
    ws.merge_cells('A5:G5')
    year_month_text = f"YEAR/MONTH: {year}/{m_name}"
    branch_text = f"BRANCH : {branch.upper()}"
    ws['A5'].value = f"{year_month_text}{' ' * 78}{branch_text}"
    ws['A5'].font = header_font
    ws['A5'].alignment = Alignment(horizontal="left", vertical="center")
    for r in range(1, 6):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c); cell.border = border; cell.fill = header_fill

    # === TABLE HEADERS ===
    headers_config = [("S.NO", 'A6:A7'), ("DATE", 'B6:B7'), (f"VC NO:\nAMIT_{b_code}_PVC_NO.", 'C6:C7'), ("PURPOSE", 'D6:E7'), ("PAYMENT AMOUNT", 'F6:G6')]
    for text, merge_range in headers_config:
        ws.merge_cells(merge_range)
        cell = ws[merge_range.split(':')[0]]; cell.value = text; cell.font = header_font; cell.alignment = center_align; cell.fill = header_fill
    ws['F7'].value = "ONLINE(A/C)"; ws['F7'].font = header_font; ws['F7'].alignment = center_align; ws['F7'].fill = header_fill
    ws['G7'].value = "CASH"; ws['G7'].font = header_font; ws['G7'].alignment = center_align; ws['G7'].fill = header_fill
    for r in [6, 7]:
        ws.row_dimensions[r].height = 30
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c); cell.border = border; cell.font = header_font
            if not cell.fill.start_color.index or cell.fill.start_color.index == '00000000': cell.fill = header_fill

    # === DATA ROWS ===
    current_row = 8
    for sno, date, vc, purpose, online, cash in data:
        ws.row_dimensions[current_row].height = 20
        ws.cell(row=current_row, column=1, value=sno)
        ws.cell(row=current_row, column=2, value=date)
        ws.cell(row=current_row, column=3, value=vc)
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
        ws.cell(row=current_row, column=4, value=purpose)
        ws.cell(row=current_row, column=6, value=online)
        ws.cell(row=current_row, column=7, value=cash)
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col); cell.font = data_font; cell.border = border
            if col == 4: cell.alignment = left_align
            elif col in [6, 7]: cell.alignment = right_align_v
            else: cell.alignment = center_align
        current_row += 1

    # === FOOTER SECTION ===
    # Row: TOTAL AMOUNT
    ws.row_dimensions[current_row].height = 25
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
    ws.cell(row=current_row, column=4, value="TOTAL AMOUNT").font = data_font_bold
    ws.cell(row=current_row, column=4).alignment = right_align_v
    ws.cell(row=current_row, column=6, value=total_online).font = data_font_bold
    ws.cell(row=current_row, column=6).alignment = right_align_v
    ws.cell(row=current_row, column=7, value=total_cash).font = data_font_bold
    ws.cell(row=current_row, column=7).alignment = right_align_v
    for col in range(1, 8): ws.cell(row=current_row, column=col).border = border
    current_row += 1

    # Row: MONTH TOTAL EXPENSE
    ws.row_dimensions[current_row].height = 25
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    ws.cell(row=current_row, column=1, value=f"{m_name} MONTH TOTAL EXPENSE").font = data_font_bold
    ws.cell(row=current_row, column=1).alignment = right_align_v
    ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=6, value=total_online + total_cash).font = data_font_bold
    ws.cell(row=current_row, column=6).alignment = center_align
    for col in range(1, 8): ws.cell(row=current_row, column=col).border = border
    current_row += 1

    # Row: SIGNATURE LABELS
    ws.row_dimensions[current_row].height = 25
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.cell(row=current_row, column=1, value="PREPARED BY").font = data_font_bold
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
    ws.cell(row=current_row, column=4, value="RECHECKED BY").font = data_font_bold
    ws.cell(row=current_row, column=6, value="APPROVED BY").font = data_font_bold
    ws.cell(row=current_row, column=7, value="VERIFIED BY").font = data_font_bold
    for col in range(1, 8):
        cell = ws.cell(row=current_row, column=col); cell.border = border; cell.alignment = center_align
    current_row += 1

    # Row: SIGNATURE SPACE
    ws.row_dimensions[current_row].height = 60
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)
    for col in range(1, 8): ws.cell(row=current_row, column=col).border = border
    current_row += 1

    # === COLUMN WIDTHS ===
    col_widths = [8, 14, 23, 80, 20, 18, 18]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # === RESPONSE ===
    final_f = f"AMIT_{b_code}_PV_{m_short}'{y_short}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{final_f}"'
    wb.save(response)
    return response


def registration_work(request, branch):
    return render(request, 'logins/registration_work.html', {'branch': branch})

product_model_map = {
    "NAGERCOIL": NagercoilProductRegistration,
    "TIRUNELVELI": TirunelveliProductRegistration,
    "PUDUKOTTAI": PudukottaiProductRegistration,
    "CHENNAI": ChennaiProductRegistration,
}

product_form_map = {
    "NAGERCOIL": NagercoilProductRegistrationForm,
    "TIRUNELVELI": TirunelveliProductRegistrationForm,
    "PUDUKOTTAI": PudukottaiProductRegistrationForm,
    "CHENNAI": ChennaiProductRegistrationForm,
}

# product Registration Report View
@login_required(login_url='login')
def product_registration_report(request, branch):
    if branch not in product_model_map:
        return redirect('dashboard')

    product_records = product_model_map[branch].objects.all().order_by('REG_CODE','S_NO','DATE')
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]
    return render(request, "logins/product_registration_report.html", {
        "branch": branch,
        "product_records": product_records,
        'years': years,
    })

# Add View
@login_required(login_url='login')
def add_product_registration(request, branch):
    form_class = product_form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('product_registration_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_product_registration.html", {
        "form": form,
        "branch_name": branch
    })

# Edit View
@login_required(login_url='login')
def edit_product_registration(request, branch, record_id):
    if branch not in product_model_map:
        return redirect('dashboard')

    model = product_model_map[branch]
    form_class = product_form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('product_registration_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_product_registration.html", {
        "form": form,
        "branch": branch
    })

# Delete View
@login_required(login_url='login')
def delete_product_registration(request, branch, record_id):
    if branch not in product_model_map:
        return redirect('dashboard')

    model = product_model_map[branch]
    record = get_object_or_404(model, id=record_id)

    # Immediately delete on GET request
    record.delete()
    messages.success(request, "product registration record deleted successfully.")
    return redirect('product_registration_report', branch=branch)

intern_model_map = {
    "NAGERCOIL": NagercoilInternshipRegistration,
    "TIRUNELVELI": TirunelveliInternshipRegistration,
    "PUDUKOTTAI": PudukottaiInternshipRegistration,
    "CHENNAI": ChennaiInternshipRegistration,
}

intern_form_map = {
    "NAGERCOIL": NagercoilInternshipRegistrationForm,
    "TIRUNELVELI": TirunelveliInternshipRegistrationForm,
    "PUDUKOTTAI": PudukottaiInternshipRegistrationForm,
    "CHENNAI": ChennaiInternshipRegistrationForm,
}

@login_required(login_url='login') 
def internship_registration_report(request, branch):
    if branch not in intern_model_map:
        return redirect('dashboard')

    records = intern_model_map[branch].objects.all().order_by('REG_CODE','S_NO','DATE')
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]
    return render(request, "logins/internship_registration_report.html", {
        "branch": branch,
        "records": records,
        'years': years,
    })

@login_required(login_url='login') 
def add_internship_registration(request, branch):
    form_class = intern_form_map.get(branch)

    if not form_class:
        return redirect('dashboard')

    if request.method == "POST":
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('internship_registration_report', branch=branch)
    else:
        form = form_class()

    return render(request, "logins/add_internship_registration.html", {
        "form": form,
        "branch_name": branch
    })

@login_required(login_url='login') 
def edit_internship_registration(request, branch, record_id):
    if branch not in intern_model_map:
        return redirect('dashboard')

    model = intern_model_map[branch]
    form_class = intern_form_map[branch]
    record = get_object_or_404(model, id=record_id)

    if request.method == "POST":
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('internship_registration_report', branch=branch)
    else:
        form = form_class(instance=record)

    return render(request, "logins/edit_internship_registration.html", {
        "form": form,
        "branch": branch
    })

@login_required(login_url='login')
def delete_internship_registration(request, branch, record_id):
    if branch not in intern_model_map:
        return redirect('dashboard')

    model = intern_model_map[branch]
    record = get_object_or_404(model, id=record_id)

    
    record.delete()
    messages.success(request, "Internship registration record deleted successfully.")
    return redirect('internship_registration_report', branch=branch)

@login_required
def download_report(request, model, branch, filename):
    month = request.GET.get('month')
    year = request.GET.get('year')

    # === Validate ===
    if not month or not year:
        return HttpResponse("Month and Year parameters are required.", status=400)

    records = model.objects.filter(DATE__year=year, DATE__month=month).order_by('S_NO', 'DATE')
    if not records.exists():
        return HttpResponse("No records found for the selected month and year.", status=404)

    # === Prepare Data ===
    data = []
    for record in records:
        data.append([
            record.S_NO,
            record.DATE.strftime('%d-%m-%Y'),
            record.REG_CODE,
            record.NAME,
            record.DEPARTMENT,
            getattr(record, 'PHD_TYPE', '') or getattr(record, 'PROJECT_TYPE', '') or
            getattr(record, 'INTERNSHIP_TYPE', '') or getattr(record, 'PUBLICATION_TYPE', ''),
            getattr(record, 'COLLEGE_UNIVERSITY', '') or getattr(record, 'institution', ''),
            record.MOBILE_NO,
            record.EMAIL_ID if record.EMAIL_ID else '-',
            record.TOTAL_AMOUNT if record.TOTAL_AMOUNT not in [None, ''] else 0,
            record.AMOUNT_PAID if record.AMOUNT_PAID not in [None, ''] else 0,
            record.AMOUNT_BALANCE if record.AMOUNT_BALANCE not in [None, ''] else 0,
            record.STATUS
        ])

    # === Workbook Setup ===
    wb = Workbook()
    ws = wb.active
    
    
    branch_map = {
        'NAGERCOIL': 'NGL',
        'TIRUNELVELI': 'TVL',
        'CHENNAI': 'TBM',
        'PUDUKOTTAI': 'PDKT',
        'VIJAYAWADA': 'VJW'
    }
    type_map = {
        'PHD_REGISTRATION': 'PHD',
        'PROJECT_REGISTRATION': 'PROJ',
        'INTERNSHIP_REGISTRATION': 'INT'
    }
    
    b_code = branch_map.get(branch.upper(), branch[:3].upper())
    # Shorten report type for title and filename
    short_type = type_map.get(filename.replace('_REPORT', ''), filename.replace('_REPORT', ''))
    
    # Format month/year for title
    title_date = datetime(int(year), int(month), 1).strftime("%b'%y").upper()
    ws.title = f"AMIT_{b_code}_{short_type} {title_date}"

    # === Styles ===
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_font = Font(bold=True, name="Bookman Old Style", size=12)
    data_font = Font(name="Bookman Old Style", size=11)

    # === LOGO + TITLE ===
    ws.row_dimensions[1].height = 62
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 40
    ws.row_dimensions[4].height = 45

    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "logo.png")
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width, logo.height = 160, 100  # Adjusted size to center in A1:A3
        ws.add_image(logo, "A1")

    # Merge for Logo (A1:B3)
    ws.merge_cells('A1:B3')
    
    # Title (C1:M1)
    ws.merge_cells('C1:M1')
    ws['C1'].value = "ATHITH MITHRA INDUSTRIAL PVT LTD"
    ws['C1'].font = Font(bold=True, color="FF0000", size=28, name="Bookman Old Style")
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")

    # Branch (C2:E2)
    ws.merge_cells('C2:E2')
    ws['C2'].value = f"BRANCH- {branch}"
    ws['C2'].font = Font(bold=True, color="FF0000", size=18, name="Bookman Old Style")
    ws['C2'].alignment = Alignment(horizontal="center", vertical="center")

    # Date (C3:E3) - Format: MMM-YY (e.g., Nov-25)
    date_obj = datetime(int(year), int(month), 1)
    date_str = date_obj.strftime('%b-%y')
    
    ws.merge_cells('C3:E3')
    ws['C3'].value = f"{date_str}"
    ws['C3'].font = Font(bold=True, color="FF0000", size=14, name="Bookman Old Style")
    ws['C3'].alignment = Alignment(horizontal="center", vertical="center")

    # Report Details (G2:M3)
    report_type = filename.replace('_', ' ').replace('REPORT', '').strip()
    
    # Calculate Academic Year (AUG to JULY)
    # If month is August (8) or later, academic year starts this year
    # Otherwise, it started previous year
    month_int = int(month)
    year_int = int(year)
    
    if month_int >= 8:  # August or later
        start_year = year_int
        end_year = year_int + 1
    else:  # Before August
        start_year = year_int - 1
        end_year = year_int
    
    # Format: AUG'25 TO JULY'26
    academic_year = f"AUG'{str(start_year)[-2:]} TO JULY'{str(end_year)[-2:]}"
    record_count = len(data)
    details_text = f"{report_type} DETAILS {academic_year} "
    
    ws.merge_cells('F2:M3')
    ws['F2'].value = details_text
    ws['F2'].font = Font(bold=True, color="0070C0", size=20, name="Bookman Old Style")
    ws['F2'].alignment = Alignment(horizontal="center", vertical="center")
    ws['F2'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Borders
    # Logo Box
    for row in ws['A1:B3']:
        for cell in row:
            cell.border = border
            
    # Title Box
    for row in ws['C1:M1']:
        for cell in row:
            cell.border = border

    # Branch Box
    for row in ws['C2:E2']:
        for cell in row:
            cell.border = border
            
    # Date Box
    for row in ws['C3:E3']:
        for cell in row:
            cell.border = border

    # Details Box
    for row in ws['F2:M3']:
        for cell in row:
            cell.border = border


    # === TABLE HEADERS ===
    headers = [
        "S.NO", "DATE", "REG CODE", "NAME", "DEPT", "TYPE", 
        "COLLEGE/UNIVERSITY", "MOBILE NO", "EMAIL ID", 
        "TOTAL AMOUNT", "AMOUNT PAID", "AMOUNT BALANCE", "STATUS"
    ]
    start_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.border = border

    ws.row_dimensions[start_row].height = 45

    # === DATA ROWS ===
    current_row = start_row + 1
    for row_data in data:
        ws.row_dimensions[current_row].height = 50  # Set uniform height for neatness
        for col, value in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col, value=value)
            cell = ws.cell(row=current_row, column=col)
            cell.font = data_font
            cell.border = border

            # Alignment based on column type
            if col in [4, 6, 7, 9]:  # NAME, TYPE, COLLEGE, EMAIL - left align
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col in [10, 11, 12]:  # TOTAL AMOUNT, AMOUNT PAID, AMOUNT BALANCE - right align
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:  # All other columns - center align
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        current_row += 1

    # === TOTAL AMOUNT ROW ===
    total_row = current_row
    
    # Calculate totals for amount columns
    from decimal import Decimal
    total_amount_sum = sum(Decimal(row_data[9] or 0) for row_data in data)  # Column 10 (TOTAL AMOUNT)
    amount_paid_sum = sum(Decimal(row_data[10] or 0) for row_data in data)  # Column 11 (AMOUNT PAID)
    amount_balance_sum = sum(Decimal(row_data[11] or 0) for row_data in data)  # Column 12 (AMOUNT BALANCE)
    
    # Merge cells A-I for "TOTAL AMOUNT" label
    ws.merge_cells(f'A{total_row}:I{total_row}')
    ws[f'A{total_row}'].value = "TOTAL AMOUNT"
    ws[f'A{total_row}'].font = Font(bold=True, name="Bookman Old Style", size=12)
    ws[f'A{total_row}'].alignment = Alignment(horizontal="right", vertical="center")
    ws[f'A{total_row}'].border = border
    
    # Apply borders to all merged cells
    for col in range(1, 10):  # Columns A-I
        ws.cell(row=total_row, column=col).border = border
    
    # Total Amount value (Column J)
    ws.cell(row=total_row, column=10, value=total_amount_sum)
    ws.cell(row=total_row, column=10).font = Font(bold=True, name="Bookman Old Style", size=12)
    ws.cell(row=total_row, column=10).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=10).border = border
    
    # Amount Paid value (Column K)
    ws.cell(row=total_row, column=11, value=amount_paid_sum)
    ws.cell(row=total_row, column=11).font = Font(bold=True, name="Bookman Old Style", size=12)
    ws.cell(row=total_row, column=11).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=11).border = border
    
    # Amount Balance value (Column L)
    ws.cell(row=total_row, column=12, value=amount_balance_sum)
    ws.cell(row=total_row, column=12).font = Font(bold=True, name="Bookman Old Style", size=12)
    ws.cell(row=total_row, column=12).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=12).border = border
    
    # Status column (Column M) - empty but with border
    ws.cell(row=total_row, column=13).border = border
    
    ws.row_dimensions[total_row].height = 30
    current_row += 1

    # === FOOTER SECTION ===
    # First row - Labels
    footer_row = current_row
    
    # PREPARED BY (columns A-C)
    ws.merge_cells(f'A{footer_row}:C{footer_row}')
    ws[f'A{footer_row}'].value = "PREPARED BY"
    ws[f'A{footer_row}'].font = Font(bold=True, name="Bookman Old Style", size=12)
    ws[f'A{footer_row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'A{footer_row}'].border = border
    for col in range(1, 4):
        ws.cell(row=footer_row, column=col).border = border
    
    # RECHECKED BY (columns D-F)
    ws.merge_cells(f'D{footer_row}:F{footer_row}')
    ws[f'D{footer_row}'].value = "RECHECKED BY"
    ws[f'D{footer_row}'].font = Font(bold=True, name="Bookman Old Style", size=12)
    ws[f'D{footer_row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'D{footer_row}'].border = border
    for col in range(4, 7):
        ws.cell(row=footer_row, column=col).border = border
    
    # APPROVED BY (columns G-J)
    ws.merge_cells(f'G{footer_row}:J{footer_row}')
    ws[f'G{footer_row}'].value = "APPROVED BY"
    ws[f'G{footer_row}'].font = Font(bold=True, name="Bookman Old Style", size=12)
    ws[f'G{footer_row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'G{footer_row}'].border = border
    for col in range(7, 11):
        ws.cell(row=footer_row, column=col).border = border
    
    # VERIFIED BY (columns K-M)
    ws.merge_cells(f'K{footer_row}:M{footer_row}')
    ws[f'K{footer_row}'].value = "VERIFIED BY"
    ws[f'K{footer_row}'].font = Font(bold=True, name="Bookman Old Style", size=12)
    ws[f'K{footer_row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws[f'K{footer_row}'].border = border
    for col in range(11, 14):
        ws.cell(row=footer_row, column=col).border = border
    
    ws.row_dimensions[footer_row].height = 25
    current_row += 1
    
    # Second row - Empty signature spaces (for actual signatures)
    signature_space_row = current_row
    
    # Empty space for PREPARED BY signature (columns A-C)
    ws.merge_cells(f'A{signature_space_row}:C{signature_space_row}')
    ws[f'A{signature_space_row}'].border = border
    for col in range(1, 4):  # Columns A-C
        ws.cell(row=signature_space_row, column=col).border = border
    
    # Empty space for RECHECKED BY signature (columns D-F)
    ws.merge_cells(f'D{signature_space_row}:F{signature_space_row}')
    ws[f'D{signature_space_row}'].border = border
    for col in range(4, 7):  # Columns D-F
        ws.cell(row=signature_space_row, column=col).border = border
    
    # Empty space for APPROVED BY signature (columns G-J)
    ws.merge_cells(f'G{signature_space_row}:J{signature_space_row}')
    ws[f'G{signature_space_row}'].border = border
    for col in range(7, 11):  # Columns G-J
        ws.cell(row=signature_space_row, column=col).border = border
    
    # Empty space for VERIFIED BY signature (columns K-M)
    ws.merge_cells(f'K{signature_space_row}:M{signature_space_row}')
    ws[f'K{signature_space_row}'].border = border
    for col in range(11, 14):  # Columns K-M
        ws.cell(row=signature_space_row, column=col).border = border
 
    ws.row_dimensions[signature_space_row].height = 50
 
    # === COLUMN WIDTHS ===
    col_widths = [10, 16, 22, 32, 20, 26, 50, 18, 45, 18, 18, 18, 22]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # === RESPONSE ===
    final_f = f"AMIT_{b_code}_{short_type}_REGISTRATION_{title_date}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{final_f}"'
    wb.save(response)
    return response

def download_product_report(request, branch):
    model_map = {
        'NAGERCOIL': NagercoilProductRegistration,
        'TIRUNELVELI': TirunelveliProductRegistration,
        'PUDUKOTTAI': PudukottaiProductRegistration,
        'CHENNAI': ChennaiProductRegistration
    }
    model = model_map.get(branch)
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    return download_report(request, model, branch, 'PROJECT_REGISTRATION')

def download_internship_report(request, branch):
    model_map = {
        'NAGERCOIL': NagercoilInternshipRegistration,
        'TIRUNELVELI': TirunelveliInternshipRegistration,
        'PUDUKOTTAI': PudukottaiInternshipRegistration,
        'CHENNAI': ChennaiInternshipRegistration
    }
    model = model_map.get(branch)
    if not model:
        return HttpResponse("Invalid branch.", status=400)
    return download_report(request, model, branch, 'INTERNSHIP_REGISTRATION')

BILL_CATEGORIES = {
    
    'product': [ 'PRODUCT','INTERNSHIP',],
    
}
def select_work_type(request, branch):
    return render(request, 'logins/select_work_type.html', {'branch': branch})

def get_model(branch, bill_type):
    model_name = f"{branch}{bill_type}BILL"
    try:
        return globals()[model_name]
    except KeyError:
        raise ValueError(f"Model '{model_name}' not found.")


def get_form(branch, bill_type):
    form_name = f"{branch}{bill_type}BILLForm"
    try:
        return globals()[form_name]
    except KeyError:
        raise ValueError(f"Form '{form_name}' not found.")


def billwise_dashboard(request, branch, work_type):
    categories = BILL_CATEGORIES.get(work_type.lower(), [])
    return render(request, 'billwise/dashboard.html', {
        'categories': categories,
        'branch': branch,
        'work_type': work_type
    })


def bill_list(request, branch, bill_type):
    try:
        model = get_model(branch, bill_type)
    except ValueError as e:
        return render(request, 'error.html', {'error': str(e)})

    records = model.objects.all().order_by('-DATE', '-S_NO')
    current_year = datetime.now().year
    years = [year for year in range(current_year, current_year - 10, -1)]

    return render(request, 'billwise/report.html', {
        'records': records,
        'branch': branch,
        'bill_type': bill_type,
        'years': years
    })

def add_bill(request, branch, bill_type):
    form_class = get_form(branch, bill_type)
    if request.method == 'POST':
        form = form_class(request.POST)
        if form.is_valid():
            form.save()
            return redirect('bill_list', branch=branch, bill_type=bill_type)
    else:
        form = form_class()
    return render(request, 'billwise/add_edit.html', {'form': form, 'branch': branch, 'bill_type': bill_type, 'action': 'Add'})

def edit_bill(request, branch, bill_type, record_id):
    model = get_model(branch, bill_type)
    record = get_object_or_404(model, id=record_id)
    form_class = get_form(branch, bill_type)
    if request.method == 'POST':
        form = form_class(request.POST, instance=record)
        if form.is_valid():
            form.save()
            return redirect('bill_list', branch=branch, bill_type=bill_type)
    else:
        form = form_class(instance=record)
    return render(request, 'billwise/add_edit.html', {'form': form, 'branch': branch, 'bill_type': bill_type, 'action': 'Edit'})

def delete_bill(request, branch, bill_type, record_id):
    model = get_model(branch, bill_type)
    record = get_object_or_404(model, id=record_id)
    record.delete()
    return redirect('bill_list', branch=branch, bill_type=bill_type)



@login_required(login_url='login')
def export_bills_to_excel(request, branch, bill_type):
    model = get_model(branch, bill_type)

    
    month = request.GET.get('month')
    year = request.GET.get('year')

    
    queryset = model.objects.all()
    if month and year:
        queryset = queryset.filter(DATE__month=month, DATE__year=year)

    
    queryset = queryset.order_by("S_NO")

    if not queryset.exists():
        return HttpResponse("No records found.", status=404)

    # === Workbook Setup ===
    wb = Workbook()
    ws = wb.active
    
    # Dynamic Sheet Title: JAN'26 BILLWISE REPORT
    month_abbr = datetime.strptime(month, "%m").strftime("%b").upper() if month and month.isdigit() else "ALL"
    year_abbr = str(year)[-2:] if year else "XX"
    ws.title = f"{month_abbr}'{year_abbr} BILLWISE REPORT"

    # === Styles ===
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_font = Font(bold=True, name="Bookman Old Style", size=12)
    data_font = Font(name="Bookman Old Style", size=11)
    
    # Custom Fonts
    company_font = Font(bold=True, color="FF0000", size=24, name="Bookman Old Style")
    details_font = Font(bold=True, color="0070C0", size=18, name="Bookman Old Style")
    branch_font = Font(bold=True, color="00B050", size=14, name="Bookman Old Style")
    month_year_label_font = Font(bold=True, color="7030A0", size=16, name="Bookman Old Style")
    month_year_value_font = Font(bold=True, color="C0504D", size=16, name="Bookman Old Style")

    # === HEADER LAYOUT ===
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 50

    # 1. Logo Area (A1:B3)
    ws.merge_cells('A1:B3')
    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "logo.png")
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width, logo.height = 120, 100  # Wider for A1:B3
        ws.add_image(logo, "A1")

    # 2. Company Name (C1:K1)
    ws.merge_cells('C1:K1')
    ws['C1'].value = "ATHITH MITHRA INDUSTRIAL PVT LTD"
    ws['C1'].font = company_font
    ws['C1'].alignment = Alignment(horizontal="center", vertical="center")

    # 3. Branch Name (C2:D3)
    ws.merge_cells('C2:D3')
    ws['C2'].value = f"BRANCH- {branch.upper()}"
    ws['C2'].font = branch_font
    ws['C2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 4. Payment Details (E2:I3)
    ws.merge_cells('E2:I3')
    display_bill_type = "PROJECT CENTRE" if bill_type.upper() == "CENTRE" else bill_type.upper()
    ws['E2'].value = f"{display_bill_type} BILLWISE PAYMENT DETAILS"
    ws['E2'].font = details_font
    ws['E2'].alignment = Alignment(horizontal="center", vertical="center")

    # 5. Month & Year Label (J2:K2)
    ws.merge_cells('J2:K2')
    ws['J2'].value = "MONTH & YEAR"
    ws['J2'].font = month_year_label_font
    ws['J2'].alignment = Alignment(horizontal="center", vertical="center")

    # 6. Month & Year Value (J3:K3)
    ws.merge_cells('J3:K3')
    month_name = datetime.strptime(month, "%m").strftime("%b").upper() if month and month.isdigit() else "ALL"
    ws['J3'].value = f"{month_name} & {year or 'ALL'}"
    ws['J3'].font = month_year_value_font
    ws['J3'].alignment = Alignment(horizontal="center", vertical="center")

    # Apply borders to header area (Rows 1-3, Columns A-K)
    for r in range(1, 4):
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = border

    # === TABLE HEADERS (Row 4) ===
    headers = [
        'S.NO', 'DATE', 'BILL NUMBER', 'REGISTRATION NUMBER', 'NAME',
        'TOTAL AMOUNT', 'CASH PAYMENT', 'ONLINE PAYMENT',
        'TOTAL PAID AMOUNT TILL NOW', 'BALANCE', 'PAYMENT STATUS'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
        cell.border = border

    # === DATA ROWS ===
    current_row = 5
    current_week = None
    week_cash = 0
    week_online = 0
    
    # Grand Totals
    grand_billed = 0
    grand_cash = 0
    grand_online = 0
    grand_paid_till_now = 0
    grand_balance = 0
    
    # Calculate first day weekday for the month (Sun=0, Mon=1, ..., Sat=6)
    if queryset.exists() and queryset[0].DATE:
        first_of_month = queryset[0].DATE.replace(day=1)
        first_weekday = (first_of_month.weekday() + 1) % 7
    else:
        first_weekday = 0

    subtotal_font = Font(bold=True, name="Bookman Old Style", size=11)
    subtotal_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light Yellow

    for record in queryset:
        # Determine Calendar Week of the Month (Sunday-start)
        day = record.DATE.day if record.DATE else 1
        week_num = (day + first_weekday - 1) // 7 + 1

        # Insert Week Subtotal and New Week Header if it changes
        if week_num != current_week:
            # 1. Insert Subtotal for the PREVIOUS week
            if current_week is not None:
                ws.cell(row=current_row, column=6, value=f"WEEK {current_week} TOTAL").font = subtotal_font
                ws.cell(row=current_row, column=6).alignment = Alignment(horizontal="center", vertical="center")
                
                # Separate Cash and Online payment totals
                ws.cell(row=current_row, column=7, value=week_cash).font = subtotal_font
                ws.cell(row=current_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(row=current_row, column=8, value=week_online).font = subtotal_font
                ws.cell(row=current_row, column=8).alignment = Alignment(horizontal="right", vertical="center")
                
                # Apply fill and borders to the subtotal row (A-K)
                for c in range(1, 12):
                    cell = ws.cell(row=current_row, column=c)
                    cell.fill = subtotal_fill
                    cell.border = border
                
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
                # Reset weekly sums
                week_cash = 0
                week_online = 0

            # 2. Insert NEW Week Header
            current_week = week_num
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
            week_cell = ws.cell(row=current_row, column=1, value=f"--- WEEK {current_week} ---")
            week_cell.font = Font(bold=True, name="Bookman Old Style", size=12, color="0000FF") # Blue bold
            week_cell.alignment = Alignment(horizontal="center", vertical="center")
            week_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light gray
            
            # Apply borders to the merged week row
            for c in range(1, 12):
                ws.cell(row=current_row, column=c).border = border
                
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # Prepare record data
        row_data = [
            record.S_NO,
            record.DATE.strftime("%d-%m-%Y") if record.DATE else "-",
            record.BILL_NUMBER,
            record.REGISTRATION_NUMBER,
            record.NAME,
            float(record.TOTAL_AMOUNT) if record.TOTAL_AMOUNT else 0,
            float(record.CASH_RECEIVED) if record.CASH_RECEIVED else 0,
            float(record.ONLINE_RECEIVED) if record.ONLINE_RECEIVED else 0,
            float(record.TOTAL_PAID_AMOUNT_TILL_NOW) if record.TOTAL_PAID_AMOUNT_TILL_NOW else 0,
            float(record.BALANCE) if record.BALANCE else 0,
            record.PAYMENT_STATUS,
        ]

        # Update weekly sums
        week_cash += float(record.CASH_RECEIVED or 0)
        week_online += float(record.ONLINE_RECEIVED or 0)

        # Update grand totals
        grand_billed += float(record.TOTAL_AMOUNT or 0)
        grand_cash += float(record.CASH_RECEIVED or 0)
        grand_online += float(record.ONLINE_RECEIVED or 0)
        grand_paid_till_now += float(record.TOTAL_PAID_AMOUNT_TILL_NOW or 0)
        grand_balance += float(record.BALANCE or 0)

        # Write record data
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            cell.border = border

            # Alignment rules
            if col in [1, 2, 3, 4]:  # S.NO, DATE, BILL NUMBER, REG NO
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col in [6, 7, 8, 9, 10]:  # Amounts
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col == 5:  # NAME
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col == 11:  # STATUS
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Insert Subtotal for the LAST week
    if current_week is not None:
        ws.cell(row=current_row, column=6, value=f"WEEK {current_week} TOTAL").font = subtotal_font
        ws.cell(row=current_row, column=6).alignment = Alignment(horizontal="center", vertical="center")
        
        # Separate Cash and Online payment totals
        ws.cell(row=current_row, column=7, value=week_cash).font = subtotal_font
        ws.cell(row=current_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=current_row, column=8, value=week_online).font = subtotal_font
        ws.cell(row=current_row, column=8).alignment = Alignment(horizontal="right", vertical="center")
        
        # Apply fill and borders to the subtotal row (A-K)
        for c in range(1, 12):
            cell = ws.cell(row=current_row, column=c)
            cell.fill = subtotal_fill
            cell.border = border
        
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # === FOOTER SECTION ===
    # First row - Labels
    footer_row = current_row
    
    # Define labels and their column ranges (A-K = 11 columns)
    footer_configs = [
        ("PREPARED BY", 1, 3),    # A-B
        ("RECHECKED BY", 4, 5),   # C-E
        ("APPROVED BY", 6, 8),    # F-H
        ("VERIFIED BY", 9, 11),   # I-K
    ]
    
    for label, start_col, end_col in footer_configs:
        ws.merge_cells(start_row=footer_row, start_column=start_col, end_row=footer_row, end_column=end_col)
        cell = ws.cell(row=footer_row, column=start_col, value=label)
        cell.font = Font(bold=True, name="Bookman Old Style", size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        
        # Apply borders to all columns in the merged range
        for c in range(start_col, end_col + 1):
            ws.cell(row=footer_row, column=c).border = border

    ws.row_dimensions[footer_row].height = 25
    current_row += 1
    
    # Second row - Signature Space
    signature_row = current_row
    for _, start_col, end_col in footer_configs:
        ws.merge_cells(start_row=signature_row, start_column=start_col, end_row=signature_row, end_column=end_col)
        
        # Apply borders to all columns in the merged range
        for c in range(start_col, end_col + 1):
            ws.cell(row=signature_row, column=c).border = border
            
    ws.row_dimensions[signature_row].height = 45

    # === COLUMN WIDTHS ===
    col_widths = [8, 15, 18, 25, 45, 18, 18, 18, 25, 18, 22]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # === RESPONSE ===
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # Branch Code Mapping
    branch_map = {
        "NAGERCOIL": "NGL",
        "TIRUNELVELI": "TVL",
        "PUDUKOTTAI": "PDKT",
        "CHENNAI": "TBM"
    }
    branch_code = branch_map.get(branch.upper(), branch.upper())
    
    # Month Name
    month_display = "ALL"
    if month and month.isdigit():
        try:
            month_display = datetime.strptime(month, "%m").strftime("%b").upper()
        except ValueError:
            pass
            
    year_display = year if year else "ALL"
    
    bill_type_abbr = {
        'JOURNAL': 'JNL', 'SHARING': 'SB', 'PATENT': 'PAT',
        'INTERNSHIP': 'INT', 'PROJECT': 'PROJ', 'CENTRE': 'PC'
    }
    bt_code = bill_type_abbr.get(bill_type.upper(), bill_type.upper())
    y_short = str(year)[-2:] if year else "ALL"
    month_year = f"{month_display}'{y_short}" if month_display != "ALL" else "ALL"
    filename = f"AMIT_{branch_code}_{bt_code}_BILLWISE_PAYMENT_{month_year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response